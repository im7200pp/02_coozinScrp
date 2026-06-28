import os
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def save_rankings_data(new_df, output_dir="data"):
    """
    Combines new rankings data with existing data, removes duplicates,
    sorts, and saves as both CSV and styled Excel files.
    
    Returns:
        combined_df (pd.DataFrame): The merged and sorted DataFrame.
    """
    os.makedirs(output_dir, exist_ok=True)
    csv_path = os.path.join(output_dir, "rankings.csv")
    excel_path = os.path.join(output_dir, "rankings.xlsx")
    
    # Ensure ID columns are string
    if "product_id" in new_df.columns:
        new_df["product_id"] = new_df["product_id"].astype(str)
        
    # Read existing file or create new DataFrame
    if os.path.exists(csv_path):
        try:
            existing_df = pd.read_csv(csv_path, dtype={"product_id": str})
            # Combine
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        except Exception as e:
            print(f"[Warning] Failed to read existing CSV file: {e}. Starting fresh.")
            combined_df = new_df
    else:
        combined_df = new_df
        
    # Drop duplicates: keep latest run for a given date, product_id, and keyword
    combined_df.drop_duplicates(subset=["date", "product_id", "keyword"], keep="last", inplace=True)
    
    # Sort values: dates ascending, products alphabetical, keywords alphabetical
    combined_df.sort_values(by=["date", "product_name", "keyword"], ascending=[True, True, True], inplace=True)
    
    # Save raw data to CSV (UTF-8 with BOM for Excel compatibility)
    combined_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    
    # Generate the pivot matrix Excel file
    generate_excel_file(combined_df, excel_path)
    
    return combined_df

def generate_excel_file(df, excel_path):
    """
    Creates a styled Excel file with two sheets:
    1. '일별 순위 현황' (Daily Rank Matrix)
    2. '상세 데이터(Raw)' (Raw data log)
    """
    # Create rank & change presentation
    def make_cell_value(row):
        rank = str(row['rank']) if pd.notna(row['rank']) else '-'
        if rank == 'nan' or not rank:
            rank = '-'
            
        change_dir = str(row['change_direction']) if 'change_direction' in row and pd.notna(row['change_direction']) else 'none'
        try:
            change_val = int(row['change_value']) if 'change_value' in row and pd.notna(row['change_value']) else 0
        except:
            change_val = 0
            
        if change_dir == 'up' and change_val > 0:
            return f"{rank} (▲{change_val})"
        elif change_dir == 'down' and change_val > 0:
            return f"{rank} (▼{change_val})"
        else:
            return rank

    pivot_df = df.copy()
    pivot_df['cell_value'] = pivot_df.apply(make_cell_value, axis=1)
    
    # Check required columns
    for col in ['mall_name', 'product_name', 'product_id', 'keyword', 'search_volume']:
        if col not in pivot_df.columns:
            pivot_df[col] = '-'
            
    # Drop duplicates for pivoting stability
    pivot_df.drop_duplicates(subset=["date", "product_id", "keyword"], keep="last", inplace=True)
    
    # Pivot to create date columns
    matrix = pivot_df.pivot(
        index=['mall_name', 'product_name', 'product_id', 'keyword', 'search_volume'],
        columns='date',
        values='cell_value'
    )
    
    # Sort columns (dates) descending so latest is first
    if not matrix.empty:
        matrix = matrix[sorted(matrix.columns, reverse=True)]
        
    matrix = matrix.reset_index()
    
    # Rename columns for presentation
    matrix.rename(columns={
        'mall_name': '쇼핑몰',
        'product_name': '상품명',
        'product_id': '상품 ID',
        'keyword': '키워드',
        'search_volume': '검색수'
    }, inplace=True)
    
    # Write Excel file
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        matrix.to_excel(writer, sheet_name='일별 순위 현황', index=False)
        
        # Clean and write Raw data
        raw_to_save = df.copy()
        raw_to_save.to_excel(writer, sheet_name='상세 데이터(Raw)', index=False)
        
    # Apply Premium Styling
    try:
        workbook = writer.book
        ws_matrix = workbook['일별 순위 현황']
        
        # Header Style: Cool Blue Gradient Feel
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Navy
        header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
        cell_font = Font(name='맑은 고딕', size=10)
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )
        
        # Style headers in Matrix sheet
        for col_idx in range(1, ws_matrix.max_column + 1):
            cell = ws_matrix.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            
        # Style data cells in Matrix sheet
        for row_idx in range(2, ws_matrix.max_row + 1):
            for col_idx in range(1, ws_matrix.max_column + 1):
                cell = ws_matrix.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                
                # Alignments
                if col_idx in [1, 2]: # 쇼핑몰, 상품명
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                elif col_idx in [3, 4]: # 상품 ID, 키워드
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else: # 검색수, 날짜별 순위들
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                # Highlight Up/Down rank changes
                val = str(cell.value) if cell.value is not None else ''
                if '▲' in val:
                    cell.font = Font(name='맑은 고딕', size=10, color='2563EB') # Blue for Up
                elif '▼' in val:
                    cell.font = Font(name='맑은 고딕', size=10, color='DC2626') # Red for Down
                    
        # Auto-adjust column widths for Matrix sheet
        for col in ws_matrix.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value) if cell.value is not None else ''
                val_len = sum(2 if ord(char) > 128 else 1 for char in val_str)
                if val_len > max_len:
                    max_len = val_len
            ws_matrix.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 40)
            
        # Style Raw Data sheet
        ws_raw = workbook['상세 데이터(Raw)']
        raw_header_fill = PatternFill(start_color='4B5563', end_color='4B5563', fill_type='solid') # Charcoal
        
        for col_idx in range(1, ws_raw.max_column + 1):
            cell = ws_raw.cell(row=1, column=col_idx)
            cell.fill = raw_header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        for row_idx in range(2, ws_raw.max_row + 1):
            for col_idx in range(1, ws_raw.max_column + 1):
                cell = ws_raw.cell(row=row_idx, column=col_idx)
                cell.font = cell_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
        # Auto-adjust widths for Raw Data sheet
        for col in ws_raw.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value) if cell.value is not None else ''
                val_len = sum(2 if ord(char) > 128 else 1 for char in val_str)
                if val_len > max_len:
                    max_len = val_len
            ws_raw.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 45)
            
        workbook.save(excel_path)
    except Exception as e:
        print(f"[Warning] Error during Excel styling: {e}")
